from io import BytesIO

import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils.db import get_cursor, close_connection
from utils.dataframe_helper import fetch_df

# --- NEW: shared payment-method label lookup, reused by the preview,
# Excel, and CSV generators below so a raw snake_case pay_method code
# (upi, net_banking, credit_card, debit_card) always renders the same
# nicely-cased label everywhere, matching the labels already used on
# the filter_transaction.html / monthly_transaction.html pages.
_PAYMENT_LABELS = {
    "upi": "UPI",
    "cash": "Cash",
    "net_banking": "Net Banking",
    "credit_card": "Credit Card",
    "debit_card": "Debit Card",
}


def _payment_label(raw_value):
    """--- NEW: small helper - returns the nice label for a pay_method
    code, or a Title Case fallback for anything not in the map, or
    '—' if the value is missing/NaN."""
    if pd.isna(raw_value) or raw_value in (None, ""):
        return "—"
    return _PAYMENT_LABELS.get(str(raw_value).lower(), str(raw_value).title())


def _blank_if_na(value, default=""):
    """--- NEW: small helper - pandas NaN is truthy in plain Python
    (`nan or default` would NOT fall back to `default`), so this uses
    pd.isna() explicitly to safely turn NaN/None into a clean default
    value for display (used for note/subcategory_name below)."""
    return default if pd.isna(value) else value


def get_transaction_history(user_id):
    """
    Fetch all transactions for a user ordered by date descending.
    """
    # --- NOT CHANGED: this still feeds templates/transaction_detail.html,
    # which is outside the scope of this request and still expects plain
    # positional tuples in this exact column order. Left untouched.

    cursor, db = get_cursor()
    if cursor is None:
        return []

    try:
        cursor.execute(
            """
            SELECT e.id, e.exp_date, a.name AS account_name, c.name AS category_name,
                   sc.name AS subcategory_name, e.amount, e.note, e.pay_method
            FROM expenses e
            LEFT JOIN accounts a ON e.account_id = a.id
            LEFT JOIN categories c ON e.category_id = c.id
            LEFT JOIN subcategories sc ON e.subcategory_id = sc.id
            WHERE e.user_id = %s AND e.type = 'expense'
            ORDER BY e.exp_date DESC, e.id DESC
            """,
            (user_id,)
        )

        return cursor.fetchall()

    finally:
        close_connection(cursor, db)


def filter_transactions(
    user_id,
    category=None,
    subcategory=None,
    account=None,
    payment=None,
    month=None,
    search=None,
    date_from=None,
    date_to=None
):
    """
    Filter transactions by category, subcategory, account, payment
    method, month, a free-text search term, and/or a date range.
    """
    # --- NOT CHANGED: this is the Filter Transactions page's service
    # function, unrelated to the Download Report backend covered by
    # this request, so it is left exactly as it was.
    cursor, db = get_cursor(dictionary=True)

    if cursor is None:
        return []

    try:
        if not any([category, subcategory, account, payment, month, search, date_from, date_to]):
            return []

        query = """
        SELECT
            e.id,
            e.exp_date,
            COALESCE(a.name, 'Unknown') AS account_name,
            COALESCE(c.name, 'Other') AS category_name,
            sc.name AS subcategory_name,
            e.amount,
            e.note,
            e.pay_method
        FROM expenses e
        LEFT JOIN accounts a ON e.account_id = a.id
        LEFT JOIN categories c ON e.category_id = c.id
        LEFT JOIN subcategories sc ON e.subcategory_id = sc.id
        WHERE e.user_id = %s AND e.type = 'expense'
        """

        params = [user_id]

        if category and category != "all":
            query += " AND c.name = %s"
            params.append(category)

        if payment and payment != "all":
            query += " AND e.pay_method = %s"
            params.append(payment)

        if month and month != "all":
            query += " AND MONTH(e.exp_date) = %s"
            params.append(int(month))

        if subcategory and subcategory != "all":
            query += " AND sc.name = %s"
            params.append(subcategory)

        if account and account != "all":
            query += " AND a.name = %s"
            params.append(account)

        if date_from:
            query += " AND e.exp_date >= %s"
            params.append(date_from)

        if date_to:
            query += " AND e.exp_date <= %s"
            params.append(date_to)

        if search:
            query += """
            AND (
                e.note LIKE %s
                OR c.name LIKE %s
                OR sc.name LIKE %s
                OR a.name LIKE %s
            )
            """
            like_term = f"%{search}%"
            params.extend([like_term, like_term, like_term, like_term])

        query += " ORDER BY e.exp_date DESC, e.id DESC"

        cursor.execute(query, tuple(params))

        return cursor.fetchall()

    finally:
        close_connection(cursor, db)


def get_user_accounts(user_id):
    """
    Returns the distinct accounts this user has actually recorded
    expenses against, e.g. [{"id": 1, "name": "Axis Bank"}, ...].
    """
    # --- NOT CHANGED: unrelated to this request, left exactly as it was.
    cursor, db = get_cursor(dictionary=True)

    if cursor is None:
        return []

    try:
        cursor.execute(
            """
            SELECT DISTINCT a.id, a.name
            FROM expenses e
            JOIN accounts a ON e.account_id = a.id
            WHERE e.user_id = %s AND e.type = 'expense'
            ORDER BY a.name
            """,
            (user_id,)
        )
        return cursor.fetchall()
    finally:
        close_connection(cursor, db)


def monthly_transaction_details(user_id):
    """
    Return a dictionary:
    {
        category_name: [
            {
                amount,
                note,
                pay_method,
                exp_date,
                account,
                subcategory
            }
        ]
    }

    for the current month.
    """
    # --- NOT CHANGED: unrelated to this request, left exactly as it was.

    now = datetime.now()
    df = fetch_df(
        """
        SELECT
            e.amount,
            e.note,
            e.pay_method,
            e.exp_date,
            COALESCE(c.name, 'Other') AS category,
            COALESCE(a.name, 'Unknown') AS account,
            sc.name AS subcategory
        FROM expenses e
        LEFT JOIN categories c
            ON e.category_id = c.id
        LEFT JOIN accounts a
            ON e.account_id = a.id
        LEFT JOIN subcategories sc
            ON e.subcategory_id = sc.id
        WHERE e.user_id = %s
        AND e.type = 'expense'
        """,
        (user_id,)
    )

    if df.empty:
        return {}

    df["exp_date"] = pd.to_datetime(df["exp_date"])

    monthly_df = df[
        (df["exp_date"].dt.month == now.month)
        &
        (df["exp_date"].dt.year == now.year)
    ]

    if monthly_df.empty:
        return {}

    category_details = (
        monthly_df
        .groupby("category")
        .apply(
            lambda x: x[
            [
                "amount",
                "note",
                "pay_method",
                "exp_date",
                "account",
                "subcategory"
            ]
        ].to_dict("records")
        )
        .to_dict()
    )

    return category_details


# ============================================================
# NEW: Download Report backend
# ============================================================

def _fetch_filtered_report_df(
    user_id,
    month=None,
    year=None,
    category=None,
    subcategory=None,
    account=None,
    payment=None
):
    """
    --- NEW: single shared query/DataFrame used by get_report_preview(),
    generate_excel_report(), and generate_csv_report() ("If possible
    create one reusable query/helper used by: preview, excel, csv").
    All three functions call this exactly once and derive whatever they
    need (totals, top-N, full export rows) from the same DataFrame,
    rather than each running their own SQL - this is also what keeps
    the filtering behavior identical across the preview and both
    export formats.

    Joins accounts/categories/subcategories exactly like the rest of
    this file's functions (e.g. filter_transactions). Every filter is
    optional - month, year, category, subcategory, account, payment -
    and a value of "all" is treated the same as "not provided", to
    match the <select> options already used in download_report.html.

    Returns a DataFrame with columns:
        exp_date, account_name, category_name, subcategory_name,
        amount, pay_method, note
    `subcategory_name` is intentionally left NULL-able (no COALESCE) -
    per the requirement, a missing subcategory should come back as
    NULL here so the caller/template can render "—" for it.
    """
    query = """
        SELECT
            e.exp_date,
            COALESCE(a.name, 'Unknown') AS account_name,
            COALESCE(c.name, 'Other') AS category_name,
            sc.name AS subcategory_name,
            e.amount,
            e.pay_method,
            e.note
        FROM expenses e
        LEFT JOIN accounts a ON e.account_id = a.id
        LEFT JOIN categories c ON e.category_id = c.id
        LEFT JOIN subcategories sc ON e.subcategory_id = sc.id
        WHERE e.user_id = %s AND e.type = 'expense'
    """
    params = [user_id]

    if month and str(month) != "all":
        query += " AND MONTH(e.exp_date) = %s"
        params.append(int(month))

    if year and str(year) != "all":
        query += " AND YEAR(e.exp_date) = %s"
        params.append(int(year))

    if category and category != "all":
        query += " AND c.name = %s"
        params.append(category)

    # Subcategory filter is matched against the subcategory actually
    # linked to each expense (via the LEFT JOIN above), which is what
    # keeps it automatically scoped "only within the selected category"
    # - an expense's sc.name can only ever be the subcategory that was
    # really chosen for it, under whatever category it really belongs
    # to, regardless of name collisions with subcategories of other
    # categories.
    if subcategory and subcategory != "all":
        query += " AND sc.name = %s"
        params.append(subcategory)

    if account and account != "all":
        query += " AND a.name = %s"
        params.append(account)

    if payment and payment != "all":
        query += " AND e.pay_method = %s"
        params.append(payment)

    query += " ORDER BY e.exp_date DESC"

    return fetch_df(query, tuple(params))


def get_report_preview(
    user_id,
    month=None,
    year=None,
    category=None,
    subcategory=None,
    account=None,
    payment=None
):
    """
    --- NEW: computes the 5 values download_report.html's preview cards
    expect:
        total_transactions, total_expense, top_category, top_account,
        top_payment_method
    for the given (all-optional) filters, by reusing
    _fetch_filtered_report_df() instead of a separate query.

    Per the "Error Handling" requirement, an empty result set still
    returns a well-formed dict (0 transactions, 0.0 expense, and None
    for every "top" field) - download_report.html already renders a
    missing/None "top" field as "—".
    """
    df = _fetch_filtered_report_df(
        user_id, month=month, year=year, category=category,
        subcategory=subcategory, account=account, payment=payment
    )

    if df.empty:
        return {
            "total_transactions": 0,
            "total_expense": 0.0,
            "top_category": None,
            "top_account": None,
            "top_payment_method": None,
        }

    # Cast to a real float column up front (the same fix already
    # applied in analysis_service.py's donut-chart bug) so every
    # aggregate below is a plain float, not a decimal.Decimal.
    df["amount"] = df["amount"].astype(float)

    total_transactions = int(len(df))
    total_expense = float(df["amount"].sum())

    top_category = df.groupby("category_name")["amount"].sum().idxmax()
    top_account = df.groupby("account_name")["amount"].sum().idxmax()

    # Drop rows with a missing/blank pay_method before ranking, so an
    # absent value can never "win" the top spot.
    payment_df = df[df["pay_method"].notna() & (df["pay_method"] != "")]
    top_payment_method = (
        _payment_label(payment_df.groupby("pay_method")["amount"].sum().idxmax())
        if not payment_df.empty else None
    )

    return {
        "total_transactions": total_transactions,
        "total_expense": total_expense,
        "top_category": top_category,
        "top_account": top_account,
        "top_payment_method": top_payment_method,
    }


def generate_excel_report(
    user_id,
    month=None,
    year=None,
    category=None,
    subcategory=None,
    account=None,
    payment=None
):
    """
    --- NEW: builds an in-memory .xlsx workbook (BytesIO - never written
    to disk) for the given (all-optional) filters, reusing
    _fetch_filtered_report_df() for the data.

    Columns: Date, Account, Category, Subcategory, Amount, Payment
    Method, Note - with "Total Transactions" / "Total Amount" written
    at the bottom, per the requirements. Headers get a bold white-on-
    accent-color fill so the export looks professional rather than
    default openpyxl styling.

    Per the "Error Handling" requirement, an empty filter result still
    produces a valid, downloadable workbook with just the header row
    and zeroed totals - it never raises or returns nothing.
    """
    df = _fetch_filtered_report_df(
        user_id, month=month, year=year, category=category,
        subcategory=subcategory, account=account, payment=payment
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Account", "Category", "Subcategory", "Amount", "Payment Method", "Note"]
    ws.append(headers)

    header_fill = PatternFill(start_color="C9773A", end_color="C9773A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    total_amount = 0.0
    if not df.empty:
        df["amount"] = df["amount"].astype(float)
        for _, row in df.iterrows():
            exp_date = row["exp_date"]
            date_str = exp_date.strftime("%Y-%m-%d") if hasattr(exp_date, "strftime") else str(exp_date)
            ws.append([
                date_str,
                row["account_name"],
                row["category_name"],
                _blank_if_na(row["subcategory_name"], default="—"),
                float(row["amount"]),
                _payment_label(row["pay_method"]),
                _blank_if_na(row["note"], default=""),
            ])
        total_amount = float(df["amount"].sum())

    total_transactions = int(len(df))

    # Blank spacer row, then the two totals rows (bold labels).
    ws.append([])
    ws.append(["Total Transactions", total_transactions])
    total_tx_row = ws.max_row
    ws.append(["Total Amount", total_amount])
    total_amt_row = ws.max_row

    bold = Font(bold=True)
    ws.cell(row=total_tx_row, column=1).font = bold
    ws.cell(row=total_amt_row, column=1).font = bold

    # Reasonable column widths so the export reads cleanly rather than
    # the default cramped openpyxl sizing.
    for idx, width in enumerate([12, 18, 16, 18, 12, 16, 30], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_csv_report(
    user_id,
    month=None,
    year=None,
    category=None,
    subcategory=None,
    account=None,
    payment=None
):
    """
    --- NEW: builds the same data as generate_excel_report() (via the
    same _fetch_filtered_report_df() helper, so filtering behavior is
    identical between the two export formats) but returns it as an
    in-memory CSV buffer using pandas, per the requirements. Same
    column names/order/payment-label formatting as the Excel export,
    plus the same trailing "Total Transactions" / "Total Amount"
    summary rows.
    """
    df = _fetch_filtered_report_df(
        user_id, month=month, year=year, category=category,
        subcategory=subcategory, account=account, payment=payment
    )

    if df.empty:
        export_df = pd.DataFrame(
            columns=["Date", "Account", "Category", "Subcategory", "Amount", "Payment Method", "Note"]
        )
        total_transactions = 0
        total_amount = 0.0
    else:
        df = df.copy()
        df["amount"] = df["amount"].astype(float)
        export_df = pd.DataFrame({
            "Date": df["exp_date"].apply(lambda d: d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else d),
            "Account": df["account_name"],
            "Category": df["category_name"],
            "Subcategory": df["subcategory_name"].apply(lambda v: _blank_if_na(v, default="—")),
            "Amount": df["amount"],
            "Payment Method": df["pay_method"].apply(_payment_label),
            "Note": df["note"].apply(lambda v: _blank_if_na(v, default="")),
        })
        total_transactions = int(len(df))
        total_amount = float(df["amount"].sum())

    buffer = BytesIO()
    export_df.to_csv(buffer, index=False)

    # Trailing summary rows, matching the Excel export's footer. These
    # intentionally have fewer columns than the data rows above (just a
    # label + a value), which is a normal/expected shape for a summary
    # footer appended to a CSV export.
    buffer.write(f"\nTotal Transactions,{total_transactions}\n".encode("utf-8"))
    buffer.write(f"Total Amount,{total_amount}\n".encode("utf-8"))

    buffer.seek(0)
    return buffer